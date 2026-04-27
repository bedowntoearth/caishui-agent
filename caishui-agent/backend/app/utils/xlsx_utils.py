"""xlsx 导入导出工具"""
import io
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ===== 样式 =====
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="微软雅黑", size=11)
CELL_FONT   = Font(name="微软雅黑", size=10)
DATE_FILL   = PatternFill("solid", fgColor="D9E1F2")
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")
thin = Side(style="thin", color="B4C6E7")
BORDER      = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_header(ws, headers: List[str], col_widths: List[int]):
    """写表头并设置列宽"""
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill    = HEADER_FILL
        cell.font    = HEADER_FONT
        cell.alignment = CENTER
        cell.border  = BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22


def _style_row(ws, row_idx: int, values: List[Any], date_cols: List[int] = None):
    """写一行数据并美化"""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font      = CELL_FONT
        cell.border    = BORDER
        cell.alignment  = CENTER if col != 2 else LEFT
        if row_idx % 2 == 0:
            cell.fill = DATE_FILL
        if isinstance(val, bool):
            cell.value = "是" if val else "否"


def to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ===== 企业导出 =====
def export_companies(rows: List[Dict]) -> bytes:
    headers  = ["ID", "企业名称", "统一社会信用代码", "法定代表人", "联系人", "联系人电话",
                "所属行业", "纳税人类型", "主管税务机关", "签约日期", "到期日期", "状态"]
    col_widths = [6, 28, 20, 12, 12, 14, 16, 12, 18, 12, 12, 8]
    date_cols  = [10, 11]

    wb = Workbook()
    ws = wb.active
    ws.title = "企业信息"
    _style_header(ws, headers, col_widths)

    for i, r in enumerate(rows, 2):
        taxpayer = {"general": "一般纳税人", "small": "小规模"}.get(r.get("taxpayer_type", ""), r.get("taxpayer_type", ""))
        status   = {"active": "服务中", "expired": "已到期"}.get(r.get("status", ""), r.get("status", ""))

        def fmt(v):
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d")
            return v or ""

        _style_row(ws, i, [
            r.get("id", ""),
            r.get("name", ""),
            r.get("credit_code", ""),
            r.get("legal_person", ""),
            r.get("contact_name", ""),
            r.get("contact_phone", ""),
            r.get("industry", ""),
            taxpayer,
            r.get("tax_authority", ""),
            fmt(r.get("sign_date")),
            fmt(r.get("expire_date")),
            status,
        ], date_cols)

    ws.freeze_panes = "A2"
    return to_bytes(wb)


# ===== 风险指标导出 =====
def export_risk_indicators(rows: List[Dict]) -> bytes:
    headers  = ["企业ID", "企业名称", "统计期间", "健康度", "综合风险等级",
                "指标1-按期申报", "指标2-库存现金", "指标3-往来款",
                "指标4-未分配利润", "指标5-股东占款", "指标6-暂估入账",
                "更新时间"]
    col_widths = [8, 24, 10, 8, 10, 14, 14, 14, 14, 14, 14, 16]

    wb = Workbook()
    ws = wb.active
    ws.title = "风险指标"
    _style_header(ws, headers, col_widths)

    LEVEL_MAP = {"normal": "正常", "remind": "提醒", "warning": "预警", "major_risk": "重大风险", "major": "重大风险"}

    for i, r in enumerate(rows, 2):
        def lvl(v): return LEVEL_MAP.get(v, v or "")

        _style_row(ws, i, [
            r.get("company_id", ""),
            r.get("company_name", ""),
            r.get("period", ""),
            r.get("health_score", ""),
            lvl(r.get("overall_level", "")),
            lvl(r.get("risk_levels", {}).get("risk1", "")),
            lvl(r.get("risk_levels", {}).get("risk2", "")),
            lvl(r.get("risk_levels", {}).get("risk3", "")),
            lvl(r.get("risk_levels", {}).get("risk4", "")),
            lvl(r.get("risk_levels", {}).get("risk5", "")),
            lvl(r.get("risk_levels", {}).get("risk6", "")),
            r.get("updated_at", ""),
        ])

    ws.freeze_panes = "A2"
    return to_bytes(wb)


# ===== 系统用户导出 =====
def export_sys_users(rows: List[Dict]) -> bytes:
    headers  = ["ID", "用户名", "真实姓名", "手机号", "角色", "状态", "备注",
                "最后登录IP", "最后登录时间", "创建时间"]
    col_widths = [6, 16, 14, 14, 14, 10, 20, 14, 16, 16]
    date_cols  = [9, 10]

    wb = Workbook()
    ws = wb.active
    ws.title = "系统用户"
    _style_header(ws, headers, col_widths)

    ROLE_MAP  = {"super_admin": "超级管理员", "admin": "管理员", "operator": "运营人员", "analyst": "财税分析师"}
    STATUS_MAP = {"active": "启用", "inactive": "禁用"}

    for i, r in enumerate(rows, 2):
        def fmt(v):
            if isinstance(v, datetime):
                return v.strftime("%Y-%m-%d %H:%M")
            return v or ""

        _style_row(ws, i, [
            r.get("id", ""),
            r.get("username", ""),
            r.get("real_name", ""),
            r.get("phone", ""),
            ROLE_MAP.get(r.get("role", ""), r.get("role", "")),
            STATUS_MAP.get(r.get("status", ""), r.get("status", "")),
            r.get("remark", ""),
            r.get("last_login_ip", ""),
            fmt(r.get("last_login_at")),
            fmt(r.get("created_at")),
        ], date_cols)

    ws.freeze_panes = "A2"
    return to_bytes(wb)


# ===== 导入解析 =====
def parse_import_file(file_bytes: bytes) -> List[Dict[str, Any]]:
    """解析上传的 xlsx，返回每行字典列表"""
    buf = io.BytesIO(file_bytes)
    from openpyxl import load_workbook
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 第一行为表头，去掉 * 和空格
    headers = [str(h).strip().replace("*", "") if h is not None else "" for h in rows[0]]
    data_rows = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue  # 跳过空行
        d = {}
        for h, v in zip(headers, row):
            if h:
                d[h] = v
        data_rows.append(d)

    wb.close()
    return data_rows
